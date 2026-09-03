#!/usr/bin/env python3
"""
Magnetic Plug Inspection ingester.

Parses one (or many) "Magnetic Plug Inspection" .xlsm workbook(s) that use the
NHL TR60 dump-truck form layout and produces:

  * assets/photos/<EQUIP>_<POS>_<DATE>.<ext>   - one photo per plug position
  * data/magnetic_plug.js                       - window.CM_DATA for the dashboard

The workbook is laid out as repeating 22-row blocks. Each block is one
equipment inspection on one date, with four magnetic-plug positions in
columns E-H (4C / 4D / 4E / 4F final-drive plugs).

Block layout (relative to the block's header row `s`):
    s      header:  Date | # | Motor Hours | <pos E> | <pos F> | <pos G> | <pos H>
    s + 1  data:    <date> | <equip no.> | <motor hrs>
    s + 16 "Particle count"        (per position, cols E-H)
    s + 17 "Component operating hrs"
    s + 18 "Oil operating hrs"
    s + 19 "Grade / Severity"      (per position, cols E-H)
    s + 20 "Comment"               (per position, cols E-H)

Photos are stored as floating images anchored to the data row of each block;
the drawing XML tells us the (column, row) each photo sits on, which is how we
attribute a photo to the right equipment + position.

The ingester is idempotent and merge-friendly: run it once per inspection file
as new rounds come in, and history accumulates per equipment (keyed by
equipment + date). Photos on the network (N:) drive can be referenced instead
of extracted -- see --photo-mode below.

Usage:
    python ingest/ingest_magnetic_plug.py path/to/Magnetic_Plug_Inspection.xlsm
    python ingest/ingest_magnetic_plug.py *.xlsm            # ingest several rounds
    python ingest/ingest_magnetic_plug.py file.xlsm --photo-mode extract

Requires: openpyxl  (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import zipfile
from datetime import date, datetime
from xml.etree import ElementTree as ET

try:
    import openpyxl
    from openpyxl.utils import get_column_letter  # noqa: F401
except ImportError:
    sys.exit("openpyxl is required. Install it with:  pip install openpyxl")

# --- Repository layout -------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
PHOTO_DIR = os.path.join(REPO, "assets", "photos")
DATA_FILE = os.path.join(REPO, "data", "magnetic_plug.js")

# --- Form geometry -----------------------------------------------------------
BLOCK_HEIGHT = 22          # rows per inspection block
FIRST_BLOCK_ROW = 2        # 1-based row of the first block's header
POSITION_COLS = [5, 6, 7, 8]   # E, F, G, H  (the four plug positions)

# Offsets from a block's header row to each labelled data row.
OFF_DATA = 1
OFF_PARTICLE = 16
OFF_COMP_HRS = 17
OFF_OIL_HRS = 18
OFF_GRADE = 19
OFF_COMMENT = 20

# Canonical position keys, in column order E,F,G,H.
POSITION_KEYS = ["4C", "4D", "4E", "4F"]

# OOXML / drawingML namespaces
NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _clean(value):
    """Normalise a cell value to a trimmed string or None."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _as_iso_date(value):
    """Return an ISO date string (YYYY-MM-DD) for a date-ish cell, else None."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _short_pos_label(header_text, key):
    """Human-friendly position label from the column header, e.g.
    '4C LEFT REAR FINAL DRIVE ...' -> '4C Left Rear Final Drive'."""
    if not header_text:
        return key
    # Keep only the leading Latin portion (English) before the Cyrillic text.
    latin = re.match(r"[0-9A-Za-z ./\-]+", str(header_text).strip())
    label = latin.group(0).strip() if latin else str(header_text).strip()
    return label or key


# --- Photo anchor extraction -------------------------------------------------
def extract_photo_anchors(xlsm_path):
    """Return {(col0, row0): media_filename} for every anchored image, plus the
    raw image bytes keyed by media filename.

    col0/row0 are 0-based (as stored in the drawing XML)."""
    anchors = {}
    media = {}
    with zipfile.ZipFile(xlsm_path) as z:
        names = z.namelist()
        drawing_names = [n for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
        for dname in drawing_names:
            rels_name = f"xl/drawings/_rels/{os.path.basename(dname)}.rels"
            rid_to_media = {}
            if rels_name in names:
                rels_root = ET.fromstring(z.read(rels_name))
                for rel in rels_root.findall("rel:Relationship", NS):
                    target = rel.get("Target", "")
                    rid_to_media[rel.get("Id")] = os.path.basename(target)

            root = ET.fromstring(z.read(dname))
            for anchor in list(root):
                frm = anchor.find("xdr:from", NS)
                if frm is None:
                    continue
                col0 = int(frm.find("xdr:col", NS).text)
                row0 = int(frm.find("xdr:row", NS).text)
                blip = anchor.find(".//a:blip", NS)
                if blip is None:
                    continue
                rid = blip.get(f"{{{NS['r']}}}embed")
                media_name = rid_to_media.get(rid)
                if media_name:
                    anchors[(col0, row0)] = media_name

        # Pull the actual image bytes for every referenced media file.
        for media_name in set(anchors.values()):
            path = f"xl/media/{media_name}"
            if path in names:
                media[media_name] = z.read(path)
    return anchors, media


# --- Workbook parsing --------------------------------------------------------
def parse_workbook(xlsm_path, photo_mode, photo_root, source_name):
    """Parse a single workbook into a list of inspection records."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    ws = wb.active

    anchors, media = extract_photo_anchors(xlsm_path)
    os.makedirs(PHOTO_DIR, exist_ok=True)

    records = []
    for s in range(FIRST_BLOCK_ROW, ws.max_row + 1, BLOCK_HEIGHT):
        equip = _clean(ws.cell(row=s + OFF_DATA, column=3).value)
        # Skip the empty template block (equipment blank or literal "TK").
        if not equip or str(equip).upper() in {"TK", "№", "#"}:
            continue

        insp_date = _as_iso_date(ws.cell(row=s + OFF_DATA, column=2).value)
        motor_hrs = _clean(ws.cell(row=s + OFF_DATA, column=4).value)

        positions = []
        for idx, col in enumerate(POSITION_COLS):
            key = POSITION_KEYS[idx]
            header = _clean(ws.cell(row=s, column=col).value)
            grade = _clean(ws.cell(row=s + OFF_GRADE, column=col).value)
            particle = _clean(ws.cell(row=s + OFF_PARTICLE, column=col).value)
            comp_hrs = _clean(ws.cell(row=s + OFF_COMP_HRS, column=col).value)
            oil_hrs = _clean(ws.cell(row=s + OFF_OIL_HRS, column=col).value)
            comment = _clean(ws.cell(row=s + OFF_COMMENT, column=col).value)

            # Photo anchored to this block's data row (0-based row = s, 0-based col = col-1).
            photo_ref = None
            media_name = anchors.get((col - 1, s))
            if media_name and insp_date:
                ext = os.path.splitext(media_name)[1].lower().lstrip(".") or "jpg"
                if ext == "jpeg":
                    ext = "jpg"
                fname = f"{equip}_{key}_{insp_date}.{ext}"
                if photo_mode == "extract":
                    out_path = os.path.join(PHOTO_DIR, fname)
                    with open(out_path, "wb") as fh:
                        fh.write(media[media_name])
                    photo_ref = f"assets/photos/{fname}"
                else:  # reference an external (e.g. N: drive) location by convention
                    photo_ref = f"{photo_root.rstrip('/')}/{fname}" if photo_root else None

            positions.append({
                "key": key,
                "label": _short_pos_label(header, key),
                # The grade is 1..5 (mobile/grade.js). The workbook writes the
                # old letters; A->1, B->2, C->3, D->4, X->5, and a digit stays.
                "grade": (
                    {"A": 1, "B": 2, "C": 3, "D": 4, "X": 5}.get(str(grade).strip().upper(),
                        int(str(grade).strip()) if str(grade).strip() in ("1", "2", "3", "4", "5") else None)
                    if grade is not None else None),
                "particleCount": particle,
                "componentHours": comp_hrs,
                "oilHours": oil_hrs,
                "comment": comment,
                "photo": photo_ref,
            })

        records.append({
            "equipment": str(equip),
            "date": insp_date,
            "motorHours": motor_hrs,
            "source": source_name,
            "positions": positions,
        })
    return records


# --- Merge & write -----------------------------------------------------------
def load_existing():
    """Read any previously-written data/magnetic_plug.js back into records."""
    if not os.path.exists(DATA_FILE):
        return []
    text = open(DATA_FILE, encoding="utf-8").read()
    m = re.search(r"window\.CM_DATA\s*=\s*(\{.*\});?\s*$", text, re.DOTALL)
    if not m:
        return []
    try:
        return json.loads(m.group(1)).get("inspections", [])
    except json.JSONDecodeError:
        return []


def merge(existing, incoming):
    """Merge on (equipment, date); newest ingest wins for a given key."""
    by_key = {(r["equipment"], r["date"]): r for r in existing}
    for r in incoming:
        by_key[(r["equipment"], r["date"])] = r
    merged = list(by_key.values())
    merged.sort(key=lambda r: (r["equipment"], r["date"] or ""))
    return merged


def write_data(records):
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    equipment = sorted({r["equipment"] for r in records})
    payload = {
        "component": "Magnetic Plug",
        "assetClass": "NHL TR60 Dump Trucks",
        "generated": None,  # left null; dashboard shows file-based history, not gen time
        "equipmentList": equipment,
        "inspections": records,
    }
    body = json.dumps(payload, ensure_ascii=False, indent=2)
    header = (
        "// AUTO-GENERATED by ingest/ingest_magnetic_plug.py -- do not edit by hand.\n"
        "// Re-run the ingester to refresh. Loaded by dashboard/index.html.\n"
    )
    with open(DATA_FILE, "w", encoding="utf-8") as fh:
        fh.write(header + "window.CM_DATA = " + body + ";\n")


def main():
    ap = argparse.ArgumentParser(description="Ingest Magnetic Plug inspection workbooks.")
    ap.add_argument("files", nargs="+", help="One or more .xlsm inspection workbooks.")
    ap.add_argument("--photo-mode", choices=["extract", "reference"], default="extract",
                    help="'extract' copies embedded photos into assets/photos/ (default). "
                         "'reference' instead writes N:-drive style paths using --photo-root.")
    ap.add_argument("--photo-root", default="",
                    help="Base path/URL used with --photo-mode reference, "
                         "e.g. 'N:/Condition Monitoring/Magnetic Plug/Photos'.")
    ap.add_argument("--fresh", action="store_true",
                    help="Ignore any existing data and rebuild from the given files only.")
    args = ap.parse_args()

    incoming = []
    for path in args.files:
        if not os.path.exists(path):
            print(f"  ! skipping (not found): {path}")
            continue
        source = os.path.basename(path)
        recs = parse_workbook(path, args.photo_mode, args.photo_root, source)
        print(f"  + {source}: {len(recs)} equipment record(s)")
        incoming.extend(recs)

    existing = [] if args.fresh else load_existing()
    merged = merge(existing, incoming)
    write_data(merged)

    n_photos = len([f for f in os.listdir(PHOTO_DIR)]) if os.path.isdir(PHOTO_DIR) else 0
    print(f"\nWrote {DATA_FILE}")
    print(f"  {len(merged)} inspection record(s), "
          f"{len({r['equipment'] for r in merged})} equipment, "
          f"{n_photos} photo file(s) in assets/photos/")


if __name__ == "__main__":
    main()
