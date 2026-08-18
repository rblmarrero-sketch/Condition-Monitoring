"""Build the lubrication data-collection workbook.

Reads mobile/lube.js — the same reference the app walks — so the models, the
unit counts and the compartment lists on the sheet are exactly what a fitter
sees on the phone. Anything already sourced is pre-filled and marked, so nobody
re-types a figure that is already in.

Run: python3 docs/build-lube-workbook.py
"""
import json, os, re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.environ.get("CM_ROOT", "/home/user/Condition-Monitoring")
OUT  = os.path.join(ROOT, "docs", "Baimskaya_lube_reference_collection.xlsx")

src    = open(os.path.join(ROOT, "mobile/lube.js"), encoding="utf-8").read()
MODELS = json.loads(re.search(r"var MODELS = (\{.*?\});\n", src, re.S).group(1))
GAPS   = json.loads(re.search(r"var GAPS = (\[.*?\]);", src, re.S).group(1))
SITE   = json.loads(re.search(r"var SITE = (\{.*?\});", src, re.S).group(1))
CAT    = json.loads(re.search(r"var CATALOG = (\[.*?\]);", src, re.S).group(1))

CLASS_NAME = {
 "HT":"Rigid dump truck","AT":"Articulated truck","EXC":"Excavator","DOZ":"Track dozer",
 "LDR":"Loader","GRD":"Grader","DRB":"Blast drill","DRE":"Exploration drill",
 "HRB":"Hydraulic rock breaker","CRJ":"Mobile jaw crusher","CRC":"Mobile cone crusher",
 "SCR":"Mobile screener"}

FONT = "Arial"
INK   = Font(name=FONT, size=10)
BOLD  = Font(name=FONT, size=10, bold=True)
HEAD  = Font(name=FONT, size=9, bold=True, color="FFFFFF")
TITLE = Font(name=FONT, size=14, bold=True)
SUB   = Font(name=FONT, size=10, color="5B686F")
GIVEN = Font(name=FONT, size=10, color="0000FF")
DONE  = Font(name=FONT, size=10, color="0C8A3E")
HDRF  = PatternFill("solid", fgColor="8A4526")
FILLIN= PatternFill("solid", fgColor="FFF7CC")
LOCKED= PatternFill("solid", fgColor="F1F4F5")
HAVE  = PatternFill("solid", fgColor="E7F4EC")
thin  = Side(style="thin", color="D6DDE1")
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ── Read me ──────────────────────────────────────────────────────────────
ws = wb.active; ws.title = "Read me"; ws.sheet_view.showGridLines = False
lines = [
 ("Baimskaya — lubrication reference collection", TITLE),
 ("Карта сбора данных по смазочным материалам", SUB), ("", None),
 ("Fill in the YELLOW cells only", BOLD),
 ("Grey columns come from the app's own asset register. GREEN rows are already sourced —", INK),
 ("leave them alone unless you think they are wrong.", INK), ("", None),
 ("Four rules that decide whether the data is usable", BOLD),
 ("1.  Capacity is the REFILL quantity — what goes in at a service, with the filter changed.", INK),
 ("     Not the dry fill, not the sump volume. They differ by up to 20%, and the wrong one", INK),
 ("     over-fills a final drive and blows the seal.", INK),
 ("2.  Copy the specification VERBATIM: \"API CK-4 / Komatsu EO-DH\", not \"engine oil\".", INK),
 ("     The exact string is what gets matched against what a product claims.", INK),
 ("3.  A figure with no document and page shows as unsourced on the dashboard. That is", INK),
 ("     deliberate — an unsourced capacity is a guess somebody will later trust.", INK),
 ("4.  If a compartment is not fitted, put N in \"Fitted?\". Blank means \"nobody has looked\",", INK),
 ("     which is a different state and is counted differently.", INK), ("", None),
 ("You can start auditing before this is finished", BOLD),
 ("Recording WHAT OIL IS IN a compartment needs none of these figures. The capacity is for", INK),
 ("topping up. The app is built so field work does not wait on this spreadsheet.", INK), ("", None),
 ("Out of scope for now", BOLD),
 ("Grease points, coolant and fuel. Grease is a real programme but a separate one — it is", INK),
 ("measured in points and intervals, not litres and specifications. Say if you want it.", INK),
]
for i, (txt, f) in enumerate(lines, start=1):
    ws.cell(i, 1, txt).font = f or INK
r = len(lines) + 2
ws.cell(r, 1, "Progress — live, from the Compartments sheet").font = BOLD
for i, (lab, f) in enumerate([
    ("Rows in the workbook",              '=COUNTA(Compartments!D2:D400)'),
    ("Complete (four facts + a source)",  '=COUNTIF(Compartments!R2:R400,"complete")'),
    ("Incomplete",                        '=COUNTIF(Compartments!R2:R400,"incomplete")'),
    ("Filled in but unsourced",           '=COUNTIF(Compartments!R2:R400,"no source")'),
    ("Marked not fitted",                 '=COUNTIF(Compartments!R2:R400,"not fitted")'),
    ("Not started",                       '=COUNTIF(Compartments!R2:R400,"not started")'),
    ("Units covered by complete rows",    '=SUMPRODUCT((Compartments!R2:R400="complete")*Compartments!C2:C400)'),
]):
    ws.cell(r+1+i, 1, lab).font = INK
    c = ws.cell(r+1+i, 3, f); c.font = BOLD; c.alignment = Alignment(horizontal="right")
ws.cell(r+9, 1, "These are live formulas. They fill in when you open this in Excel or "
                "Google Sheets — a blank there is not a fault.").font = SUB
ws.column_dimensions["A"].width = 48; ws.column_dimensions["C"].width = 12

# ── Compartments ─────────────────────────────────────────────────────────
cs = wb.create_sheet("Compartments")
HDRS = ["Class","Model","Units","Code","Compartment (EN)","Compartment (RU)","Fitted?",
        "Refill L","OEM specification (verbatim)","Interval h","Grade — winter (≤ −20 °C)",
        "Grade — summer (> −20 °C)","Manual document","Page","Checked by","Date","Notes","Status"]
WID  = [7,30,7,7,32,30,8,9,34,10,20,20,26,8,14,12,30,13]
for c, h in enumerate(HDRS, start=1):
    x = cs.cell(1, c, h); x.font = HEAD; x.fill = HDRF; x.border = BOX
    x.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cs.column_dimensions[get_column_letter(c)].width = WID[c-1]
cs.row_dimensions[1].height = 40
cs.freeze_panes = "D2"

row = 2
order = sorted(MODELS.items(), key=lambda kv: (kv[1]["cls"], -kv[1]["n"], kv[1]["m"]))
for key, M in order:
    for c in M["comps"]:
        have = c.get("cap") is not None
        vals = [M["cls"], M["m"], M["n"], c["k"], c["en"], c["ru"]]
        for i, v in enumerate(vals, start=1):
            x = cs.cell(row, i, v); x.font = GIVEN if i <= 3 else INK
            x.fill = LOCKED if i <= 3 else PatternFill(); x.border = BOX
        for i in range(7, 18):
            x = cs.cell(row, i); x.fill = HAVE if have else FILLIN; x.border = BOX; x.font = INK
        if have:
            # Already sourced. Pre-filled and green so nobody re-types it — and
            # so a wrong one is visible rather than buried in a blank sheet.
            cs.cell(row, 7,  "Y").font = DONE
            cs.cell(row, 8,  c["cap"]).font = DONE
            cs.cell(row, 9,  c.get("spec","")).font = DONE
            cs.cell(row, 10, c.get("iv","")).font = DONE
            gr = c.get("gr") or []
            cold = [g["g"] for g in gr if g["lo"] <= SITE["design"]]
            warm = [g["g"] for g in gr if g["lo"] >  SITE["design"]]
            if cold: cs.cell(row, 11, cold[0]).font = DONE
            if warm: cs.cell(row, 12, warm[0]).font = DONE
            s = c.get("src")
            if s:
                doc = str(s.get("doc",""))
                cs.cell(row, 13, doc).font = DONE
                cs.cell(row, 14, "see doc").font = DONE
                cs.cell(row, 15, s.get("who","")).font = DONE
                cs.cell(row, 16, s.get("when","")).font = DONE
        if M.get("alias"):
            cs.cell(row, 17, "Register says %s; identified as %s by %s"
                    % (M["m"], M["alias"]["is"], M["alias"]["who"])).font = SUB
        cs.cell(row, 16).number_format = "yyyy-mm-dd"
        cs.cell(row, 18,
            '=IF(UPPER(G{r})="N","not fitted",'
            'IF(AND(H{r}="",I{r}="",J{r}=""),"not started",'
            'IF(OR(H{r}="",I{r}="",J{r}=""),"incomplete",'
            'IF(OR(M{r}="",N{r}=""),"no source","complete"))))'.format(r=row))
        cs.cell(row, 18).font = INK; cs.cell(row, 18).border = BOX
        cs.cell(row, 18).alignment = Alignment(horizontal="center")
        row += 1
LAST = row - 1
dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv.promptTitle = "Fitted?"; dv.prompt = "Y if this model has this compartment, N if not."
cs.add_data_validation(dv); dv.add("G2:G%d" % LAST)
cs.auto_filter.ref = "A1:R%d" % LAST

# ── Products in use ──────────────────────────────────────────────────────
ps = wb.create_sheet("Products in use")
PH = ["Product name (as printed on the label)","Grade",
      "Specifications it claims (verbatim from the PDS)","Lowest ambient it is rated to (°C)",
      "Where it is held","Tank / drum ID","Approx. stock (L)","Photo of label taken?","Notes"]
for c, h in enumerate(PH, start=1):
    x = ps.cell(1, c, h); x.font = HEAD; x.fill = HDRF; x.border = BOX
    x.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ps.column_dimensions[get_column_letter(c)].width = [40,16,44,17,24,16,14,16,30][c-1]
ps.row_dimensions[1].height = 40; ps.freeze_panes = "A2"
# The catalogue the app already knows about, so this starts as a check rather
# than a blank page. Correct it — this is what we THINK is on site.
for i, p in enumerate(CAT):
    rr = 2 + i
    for c, v in enumerate([p["p"], p["g"], p["s"], p["lo"], "", "", "", "", ""], start=1):
        x = ps.cell(rr, c, v); x.border = BOX
        x.font = GIVEN if c <= 4 else INK
        x.fill = LOCKED if c <= 4 else FILLIN
for rr in range(2 + len(CAT), 2 + len(CAT) + 30):
    for c in range(1, 10):
        x = ps.cell(rr, c); x.fill = FILLIN; x.border = BOX; x.font = INK
ps.cell(2 + len(CAT) + 32, 1,
        "Blue rows are what the app currently believes is available. Correct them, and add "
        "anything on site that is missing — including drums you suspect are wrong.").font = SUB

# ── Register gaps ────────────────────────────────────────────────────────
gs = wb.create_sheet("Register gaps"); gs.sheet_view.showGridLines = False
gs.cell(1, 1, "Units the register records by make only").font = TITLE
gs.cell(2, 1, "A capacity cannot be attached to \"KOMATSU\". Where somebody has identified the "
              "machine, the app uses those figures today — but the register is still wrong and "
              "these stay listed until the model is recorded in 1C.").font = SUB
GH = ["Class","Recorded as","Units","Identified as","By","Actual model (confirm / fill in)","Fixed in 1C?"]
for c, h in enumerate(GH, start=1):
    x = gs.cell(4, c, h); x.font = HEAD; x.fill = HDRF; x.border = BOX
    x.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    gs.column_dimensions[get_column_letter(c)].width = [8,20,8,26,16,30,14][c-1]
gr = 5
for g in GAPS:
    key = g["cls"] + "|" + g["as"]
    units = MODELS.get(key, {}).get("n", "")
    al = g.get("alias")
    for c, v in enumerate([g["cls"], g["as"], units,
                           al["is"] if al else "", al["who"] if al else ""], start=1):
        x = gs.cell(gr, c, v); x.border = BOX
        x.font = DONE if (al and c in (4, 5)) else GIVEN
        x.fill = LOCKED
    for c in (6, 7):
        x = gs.cell(gr, c); x.fill = FILLIN; x.border = BOX; x.font = INK
    gr += 1
gs.cell(gr, 2, "Total").font = BOLD
gs.cell(gr, 3, "=SUM(C5:C%d)" % (gr - 1)).font = BOLD

# ── Site ─────────────────────────────────────────────────────────────────
ss = wb.create_sheet("Site"); ss.sheet_view.showGridLines = False
ss.cell(1, 1, "Site constants").font = TITLE
ss.cell(2, 1, "Every cold-weather verdict is computed against these. Getting the design "
              "minimum wrong changes which products qualify.").font = SUB
for c, w in zip("ABC", (44, 16, 70)): ss.column_dimensions[c].width = w
for c, h in enumerate(["Item","Value","Note"], start=1):
    x = ss.cell(4, c, h); x.font = HEAD; x.fill = HDRF; x.border = BOX
for i, (lab, val, note) in enumerate([
 ("Design minimum ambient (°C)", SITE["design"],
  "The temperature the oil must still work at — usually not the record low."),
 ("Record winter low (°C)", SITE.get("winter",-45), "Reference only."),
 ("Summer high (°C)", SITE.get("summer",28), ""),
 ("Operating hours per machine per year", SITE.get("hoursPerYear",5000),
  "Drives how fast an audit goes stale."),
 ("Lube audit crews available", 2, ""),
 ("Compartments one crew can audit per day", 6, ""),
 ("Oil analysis laboratory", "", "Name, and whether they can export results as CSV."),
 ("Who signs off the site standard", "",
  "One name. A standard anyone can change is not a standard."),
]):
    rr = 5 + i
    ss.cell(rr, 1, lab).font = INK;  ss.cell(rr, 1).border = BOX
    v = ss.cell(rr, 2, val); v.font = INK; v.fill = FILLIN; v.border = BOX
    n = ss.cell(rr, 3, note); n.font = SUB; n.border = BOX
    n.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
done = sum(1 for M in MODELS.values() for c in M["comps"] if c.get("cap") is not None)
print("wrote", OUT)
print("rows:", LAST - 1, "| models:", len(MODELS),
      "| already sourced:", done, "| gaps:", len(GAPS))
