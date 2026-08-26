"""Generate mobile/lube.js from the site's own Lubrication Masterlist.

   docs/source/Lube_Matrix_Oil_Analysis_Sampling.xlsm  ->  mobile/lube.js

THE CODES AND THE FIGURES ARE THE CLIENT'S, NOT OURS.
The component codes (1, 2, 3, 3A, 4A, 4AL … 16) are already on the printed
forms in the ute, on the per-machine sampling sheets, and in the fitters'
heads. An invented scheme would have bought nothing and cost all three. Same
for the capacities and the change intervals: they came out of OEM manuals, and
anything this script "improves" is a number somebody will later have to defend.

WHAT THIS SCRIPT DOES NOT DO is decide anything. Where the masterlist is
ambiguous or self-contradictory it carries the contradiction through and flags
it, because a generator that quietly resolves the client's data is a generator
that hides the one thing they need to see.

Run: python3 docs/build-lube-data.py
"""
import json, os, re, sys, warnings, collections
warnings.filterwarnings("ignore")
from openpyxl import load_workbook

ROOT = os.environ.get("CM_ROOT", "/home/user/Condition-Monitoring")
SRC  = os.path.join(ROOT, "docs/source/Lube_Matrix_Oil_Analysis_Sampling.xlsm")
ASSETS = os.path.join(ROOT, "mobile/assets.js")
OUT  = os.path.join(ROOT, "mobile/lube.js")
REPORT = os.path.join(ROOT, "docs/lube-import-report.txt")
SHEET = "Lube Frequency BMSK"

# ── the site's products ──────────────────────────────────────────────────
# Read from row 1 of the masterlist, columns M..T. NOT from the Lube Legend
# tab: its column letters are stale — it calls column N the hydraulic oil, and
# column N is the TO-4. The sheet is the live document, so the sheet wins.
#
# TYPE is what the fitter sees on the wall. The site already colour-codes by
# type — blue engine, red hydraulic, yellow compressor, green rock-drill, cream
# grease — and that is the right choice for a poster: eight colours to learn,
# and the colour survives a change of supplier. Colour by PRODUCT would have to
# be relearned the day a drum changes.
PRODUCT_COL = list(range(13, 21))            # M..T
TYPE_BY_CODE = {
    "0W40":       ("engine",     "Engine oil",      "Моторное"),
    "5W30":       ("powertrain", "Powertrain / TO-4","Трансмиссионное TO-4"),
    "NEXVG32":    ("hydraulic",  "Hydraulic",       "Гидравлическое"),
    "75W140":     ("gear",       "Gear oil",        "Трансмиссионное"),
    "EXVG32":     ("rockdrill",  "Rock-drill oil",  "Буровое"),
    "Grease":     ("grease",     "Grease",          "Смазка"),
    "Coolant":    ("coolant",    "Coolant",         "Антифриз"),
    "Compressor": ("compressor", "Compressor oil",  "Компрессорное"),
}
# The site's own colours, taken from the fills already used on the sheet.
# Kept as data because they are the client's convention, not a design choice.
TYPE_HUE = {
    # THE SITE'S OWN COLOURS, read off the "Template indicator" and "Indicator"
    # tabs. Not approximations of them — these are the exact fills, so the
    # workshop wall, the workbook and the app agree.
    "gearbox":    "#7030A0",   # purple
    "gear":       "#7030A0",   # the masterlist's gear oil IS the gearbox colour
    "powertrain": "#00B050",   # green  — "transmission" on their sheet
    "coolant":    "#00B0F0",   # cyan
    "engine":     "#0070C0",   # blue
    "hydraulic":  "#FF0000",   # red
    "compressor": "#FFFF00",   # yellow
    # Two the workbook draws but does not fill; taken from the rendered sheet.
    "grease":     "#C4BD97",   # tan
    "torque":     "#E36C09",   # orange
    "rockdrill":  "#948A54",   # no colour assigned yet on their sheet
}

# Brand is a SECOND dimension, and on their sheet it is the cell border rather
# than the fill. That is the right way round and worth saying why: the failure
# you are preventing is engine oil in a final drive, and that destroys the final
# drive whichever brand it came out of. Consequence follows FUNCTION, so
# function gets the fill. Brand changes when a tender is won, so it gets the
# edge — and it should disappear as the site converges on one brand per
# function, which is what the standardisation programme is for.
BRAND_HUE = {
    "Exsoil": "#FFC000",   # amber   (from the workbook)
    "Lemarc": "#D24E51",   # red     (from the workbook)
    "Shell":  "#4BACC6",   # blue    (from the rendered sheet)
    "Nexxol": "#000000",   # black   (from the rendered sheet)
    "Katana": "#808080",   # grey    (from the rendered sheet)
}
def brand_of(name):
    u = str(name or "").upper()
    for b in BRAND_HUE:
        if b.upper() in u: return b
    return None

# ── which product serves which component ─────────────────────────────────
# From the Lube Legend tab's own component table, by TYPE rather than by its
# stale column letters.
LEGEND_TYPE = {
    # 3A Steering: the Legend files it under gear oil (column O). Confirmed by
    # R. Marrero that steering is hydraulic - which is what the one machine
    # carrying it, an NHL TR60 with a 72 L steering circuit, would want. Left as
    # gear oil, the app would have offered a fitter a 75W90 for a steering pump.
    "1":"engine", "2":"gear", "3":"hydraulic", "3A":"hydraulic",
    "4":"gear","4A":"gear","4B":"gear","4C":"gear","4E":"gear","4F":"gear",
    "4AL":"gear","4AR":"gear","4BL":"gear","4BR":"gear","4CL":"gear","4CR":"gear",
    "4I":"gear","4J":"gear","4K":"gear","4L":"gear","4M":"gear","4N":"gear",
    "4O":"gear","4P":"grease",
    "5":"gear","5A":"gear","5B":"gear",
    "6":"gear","6A":"gear","6B":"gear","6C":"gear","6D":"gear","6E":"grease",
    "7":"gear","7B":"gear",
    "8":"compressor", "9":"coolant", "10":None,
    "11A":"gear","11B":"gear","11C":"gear","11D":"rockdrill","11E":"grease",
    "12A":"gear","12B":"gear","12C":"gear","12D":"gear","12E":"gear",
    "12F":"gear","12G":"gear","12H":"gear",
    "13":"grease",
    "14A":"gear","14B":"gear","14C":"gear","14E":"gear",
    "15":"wirerope", "16":"opengear",
}
# Flagged, not resolved. The Legend files the transmission under gear oil, but
# the OEM codes on the sheet for those same compartments read TO10 / TO-4 SAE
# 50, and the site stocks a TO-4. Transmission and final drive are different
# friction chemistry — a wet clutch needs the TO-4 frictional properties and a
# gear set does not — so this is an engineer's decision, not an importer's.
QUESTION_TYPES = {"2"}

# ── the rotary head ────────────────────────────────────────────────────────
# Code 7 is in the Legend and no machine uses it, because the frequency sheet
# has no column for it. Confirmed by R. Marrero that the drills do have one, so
# the compartment is real and the FIGURES are what is missing - a different
# thing from the compartment not existing. A rotary head can be audited without
# knowing its capacity; it cannot be audited if it is not on the list.
#
# Which machines get one is derived rather than typed: any model the masterlist
# already gives a drill-rig drivetrain code. 11E is deliberately NOT in this
# set - a thread-lube container sits on an excavator-mounted attachment too,
# and an excavator has no rotary head.
DRILL_CODES = {"11A", "11B", "11C", "11D", "7B"}
ROTARY_HEAD = {"k": "7", "en": "Rotary Head", "ru": "\u0412\u0440\u0430\u0449\u0430\u0442\u0435\u043b\u044c",
               "t": "gear", "verify": 1}

# The Legend types a compartment by where it sits on the machine. The OEM code
# in the next column says what the maker actually wants in it, and sometimes
# the two disagree — a "gear" compartment whose OEM code is a TO-4 wet-clutch
# oil, or a Komatsu hydraulic oil. That disagreement is not something to average
# out: TO-4 friction chemistry and GL-5 EP chemistry are different oils for
# different reasons, and GL-5 in a wet brake glazes the discs. Where the two
# columns disagree the compartment becomes a question for the engineer, and
# nothing prints a product for it until a person has answered.
CONTRADICTS = {
    "gear": re.compile(r"\bTO-?\s?4\b|\bTO\s?10\b|\bTO\s?30\b|\bTOS|"
                       r"\bHO-|\bHVLP\b|\bHYDRAUL", re.I),
}

# ── model names ──────────────────────────────────────────────────────────
# One canonical name per machine, and it is the MASTERLIST's, because that is
# the name on the sampling forms. The register's spellings are aliases onto it.
# Confirmed by R. Marrero: the 27 articulated trucks filed as "KOMATSU", and
# the one filed as "KOMATSU HM400", are all HM400-3MO.
CANON = {
    "KOMATSU":            {"AT": "Komatsu HM400-3MO"},
    "KOMATSU HM400":      {"AT": "Komatsu HM400-3MO"},
}

# How a MAKE is spelled on screen.
#
# Not a matching rule — MAKE_SYN below already folds the variants for matching,
# so nothing here changes which machine an entry lands on. This is only what a
# fitter reads. The masterlist writes "Luigong" with the u and the i the wrong
# way round, and shouts "LIUGONG" in another row; because the masterlist's name
# is the canonical one, that typo was the name on the fleet matrix, on the shop
# poster and in the report.
#
# Applied to whole words only, so a model code that happens to contain the
# letters is untouched.
SPELL = {
    "LUIGONG": "LiuGong",
    "LIUGONG": "LiuGong",
}
def spell(label):
    return re.sub(r"[A-Za-z]+",
                  lambda m: SPELL.get(m.group(0).upper(), m.group(0)), label)

# Register spelling -> masterlist spelling, where the two documents name the
# same machine differently and no general rule can safely bridge them.
#
# EVERY LINE HERE IS A JUDGEMENT, not a rule, and every one is printed in
# docs/lube-import-report.txt so it can be argued with. They are here rather
# than in the matcher because a fuzzy rule loose enough to catch "D275.5D" =
# "D275A-5" is loose enough to catch things that are not the same machine, and
# this project has already put a loader on a truck's compartments once.
ALIAS = {
    "KOMATSU D275.5D":          "Komatsu D275A-5",
    "Boart Longyear LF-90D":    "Boart Longyear LF90D Drill",
    "HITACHI ZX330-5G RB":      "Hitachi ZX 330-5G",
    "HITACHI ZX470LC-5G":       "Hitachi ZX 470",
    "HITACHI ZX470LCR-5G":      "Hitachi ZX 470",
    "KOMATSU PC2000-8 BH":      "KOMATSU PC2000-8",
    "KOMATSU PC800-8E0 (SE)":   "Komatsu PC800-8EO",
    "LiuGong CLG990FHD":        "LiuGong 990FHD",
    "TLP-4M":                   "TLP-4M-030 (ТЛП-4М-030)",
    "CHSDM DZ-98V.00100-111":   "CHSDM DZ-98V",

    # Same model code on both sides; the register drops the make, or the
    # masterlist appends what the machine IS. Category agrees on every one
    # (loader/loader, crane/crane, manlift/manlift, compactor/compactor), and
    # the code is unique in both documents — so these are the same machine, not
    # a resemblance. 32 machines that had no lubrication standard now have one.
    "CLGF330":                             "LiuGong CLGF330",
    "ELCOS GE.CU.030/027.SS+011":          "Elcos Genset GE.CU.030/027.SS+011",
    "TEREX AC220-5":                       "TEREX DEMAG AC220-5 CRANE, MOBILE",
    "TEREX AC350-6":                       "TEREX DEMAG AC350-6 CRANE, MOBILE",
    "ZRT400":                              "ZOOMLION ZRT400",
    "IVECO 473916":                        "IVECO AMT 473916 EMERGENCY EQUIPMENT FIRE-TRUCK",
    "MTGE3":                               "IVECO-AMT MTGE3 Tractor Truck",
    "JCB VM200D":                          "JCB VM200D COMPACTOR, ROLLER DRUM",
    "SX-105XC":                            "TEREX GENIE SX-105XC Manlift",
    "SX-125XC":                            "TEREX GENIE SX-125XC Manlift",
    "ZRS4531":                             "ZOOMLION ZRS4531 KONTAINER LOADER",
    # Two more where the codes differ by a suffix rather than by the make. Same
    # class of judgement as the HITACHI ZX330-5G RB and KOMATSU D275.5D lines
    # above, and marked here for the same reason: somebody should confirm them.
    "ZOOMLION ZCC2600":                    "ZOOMLION ZCC2600 260 TON",
    "Sinomach TTC025G2-V":                 "SINOMACH TTC025G2 Truck Crane",
    # NOT aliased on purpose, and listed so the omission is visible:
    #   CATERPILLAR 336-07  vs  CAT 336D — a 336-07 is not a 336D.
    #   HITROCK HMB4500 / HB3500 / HB4500+ — no breaker in the masterlist yet.
}

# ── matching a masterlist name to a register name ────────────────────────
# The same machine is spelled several ways across the two documents:
# Liugong / LiuGong / Luigong, CAT / CATERPILLAR, PC800-8EO with a letter O
# against PC800-8E0 with a zero, LF90D against LF-90D, and a trailing
# description ("Dump Truck", "Excavator") on one side and not the other.
#
# Guessing across those is how a loader gets a truck's compartments, so the
# rule is narrow: strip punctuation, fold the manufacturer synonyms, drop the
# words that describe what a machine IS rather than which model it is, and
# treat letter-O and zero as the same character. What survives has to match
# EXACTLY. Anything that does not is reported, never guessed.
MAKE_SYN = {
    "CATERPILLAR": "CAT", "LUIGONG": "LIUGONG", "KOMATSU": "KOMATSU",
    "HITACHI": "HITACHI", "SHANTUI": "SHANTUI", "BOARTLONGYEAR": "LONGYEAR",
    "BOART": "", "MCCLOSKEY": "MCCLOSKEY", "NHL": "NHL",
}
DESCRIPTORS = [
    "DUMPTRUCK","TRACTORTRUCK","TRUCKDUMP","TRUCKTRACTOR","TRUCKBOOM",
    "TRUCKWATER","TRUCKTANKER","TRUCKMECHANIC","TRACKDRILL","TRACKLOADER",
    "SKIDSTEERLOADER","SKIDSTEER","TELESCOPICHANDLER","TYREHANDLER",
    "CYLINDERHANDLER","WHEELCRANE","TRUCKCRANE","CRANEMOBILE","MOBILECRANE",
    "ALLTERRAINVEHICLE","WHEELDOZER","STEAMGENERATOR","INDUSTRIALHEATER",
    "WELDINGMACHINE","KONTAINERLOADER","CONTAINERLOADER","EMERGENCYEQUIPMENT",
    "EXCAVATOR","COMPACTOR","ROLLERDRUM","MANLIFT","GENERATOR","GENSET",
    "CRANE","LOADER","DOZER","DRILL","TRUCK","PICKUP","CREWBUS","BUS",
    "FIRETRUCK","FORKLIFT","TRACTOR",
]
def _base(s):
    s = re.sub(r"[^A-Z0-9]", "", str(s or "").upper())
    for k, v in MAKE_SYN.items():
        if s.startswith(k): s = v + s[len(k):]; break
    return s
def keys(s):
    """Every normal form this name could legitimately be written as."""
    b = _base(s)
    out = {b}
    for d in DESCRIPTORS:
        if b.endswith(d) and len(b) > len(d) + 2: out.add(b[:-len(d)])
        if b.startswith(d) and len(b) > len(d) + 2: out.add(b[len(d):])
    more = set()
    for k in out:
        more.add(k.replace("O", "0"))          # PC800-8EO  vs  PC800-8E0
    out |= more
    return {k for k in out if len(k) >= 4}

def norm(s):
    return _base(s)

def read_products(ws):
    out = []
    for c in PRODUCT_COL:
        name = str(ws.cell(1, c).value or "").strip()
        code = str(ws.cell(2, c).value or "").strip()
        if not name or not code: continue
        t = TYPE_BY_CODE.get(code)
        if not t:
            print("  ! unknown product code in row 2:", code); continue
        br = brand_of(name)
        out.append({"p": name, "code": code, "t": t[0], "en": t[1], "ru": t[2],
                    "hue": TYPE_HUE.get(t[0], "#8b969c"),
                    "brand": br, "bhue": BRAND_HUE.get(br) if br else None})
    return out

# ── the Legend, which is where the curated English names live ──────────────
# The header row of the frequency sheet carries whatever was typed when the
# column was added, and for twelve components that is Russian in the English
# field, one truncated word, and one with a space broken into the middle of a
# word ("Направля ющее колесо"). The Legend tab has a proper English name for
# every one of them and nobody was reading it.
def read_legend(wb):
    if "Lube Legend" not in wb.sheetnames:
        return {}
    ws, out = wb["Lube Legend"], {}
    for r in range(1, ws.max_row + 1):
        code = str(ws.cell(r, 1).value or "").strip()
        name = str(ws.cell(r, 2).value or "").strip()
        if not code or not name:
            continue
        if not re.match(r"^\d{1,2}[A-Z]{0,2}$", code):
            continue                        # headings, colour keys, prose
        name = re.sub(r"\s*\*+\s*$", "", name)          # "newly added" marker
        name = re.sub(r"\s*\([^)]*[А-Яа-яЁё][^)]*\)", "", name)  # RU gloss
        name = re.sub(r"\s*\(unified[^)]*\)", "", name, flags=re.I)
        name = re.sub(r"\s+", " ", name).strip()
        if name:
            out[code] = name
    return out


# Fuel is code 10 and it is NOT imported. The Legend says so itself - "Fuel,
# not a lubricant, tracked for volume only" - and putting a fuel tank in front
# of a fitter doing an oil audit is how the round stops being an oil audit.
SINGLETON = {"9": "coolant"}


def read_singletons(ws):
    """Coolant is a lone capacity column with no interval and no OEM beside it,
       so the triple rule below skips it - and it has been skipping it since the
       first import. Forty-one machines carry a coolant capacity in this sheet,
       there is an antifreeze on the shelf and a hundred thousand litres of it
       on the 2027 specification, and not one compartment existed to record it
       against. Found by its own heading rather than by column letter, because
       the Legend's column letters are already three out of date."""
    out = []
    for c in range(1, ws.max_column + 1):
        code = str(ws.cell(2, c).value or "").strip()
        if code not in SINGLETON:
            continue
        name = str(ws.cell(1, c).value or "").replace("\n", " ").strip()
        if not name:
            continue
        en, ru = split_bilingual(name)
        out.append({"k": code, "en": en, "ru": ru, "col": c,
                    "t": SINGLETON[code], "single": True})
    return out


def read_components(ws):
    """Every component triple: the code is in row 2 over the CAPACITY column,
       the next column is the interval and the one after is the OEM string."""
    comps = []
    for c in range(23, ws.max_column + 1):
        code = ws.cell(2, c).value
        if code is None: continue
        name = str(ws.cell(1, c).value or "").replace("\n", " ").strip()
        if not name or name.lower().startswith("none"): continue
        if str(ws.cell(1, c + 1).value or "").strip().lower() != "frequency of replacements":
            continue                       # not a component triple
        en, ru = split_bilingual(name)
        comps.append({"k": str(code).strip(), "en": en, "ru": ru, "col": c})
    return comps

CYR = re.compile(r"[А-Яа-яЁё]")
def split_bilingual(s):
    """The headers carry both languages in one cell: "Engine/ORS Двигатель".
       Split on the first Cyrillic character so each language gets its own
       field, rather than every label being shown twice on a bilingual app."""
    s = re.sub(r"\s+", " ", s).strip()
    m = CYR.search(s)
    if not m: return s, s
    en = s[:m.start()].strip(" /–-—")
    ru = s[m.start():].strip()
    return (en or s), (ru or s)

def main():
    if not os.path.exists(SRC):
        sys.exit("masterlist not found: " + SRC)
    wbv = load_workbook(SRC, data_only=True)
    wbf = load_workbook(SRC)                      # a second pass, for the fills
    wv, wf = wbv[SHEET], wbf[SHEET]

    products = read_products(wv)
    comps    = read_components(wv) + read_singletons(wv)
    legend   = read_legend(wbv)
    # An English field with Cyrillic in it, or empty, is not an English name.
    # Those get the Legend's; anything else keeps what the sheet says and is
    # REPORTED instead, because silently rewriting a name somebody typed on
    # purpose is how a reference stops matching the paper in the ute.
    renamed, differ, rotary = [], [], []
    for cp in comps:
        want = legend.get(cp["k"])
        if not want:
            continue
        if CYR.search(cp["en"] or "") or not (cp["en"] or "").strip():
            renamed.append((cp["k"], cp["en"], want))
            cp["en"] = want
        elif cp["en"].strip().lower() != want.strip().lower():
            differ.append((cp["k"], cp["en"], want))
    notes, questions = [], []

    # the register, so masterlist rows can be tied to real machines
    raw = open(ASSETS, encoding="utf-8").read()
    assets = json.loads(re.search(r"window\.ASSETS\s*=\s*(\[.*?\]);", raw, re.S).group(1))
    reg = collections.defaultdict(set)            # normal form -> {(cls, model)}
    counts = collections.Counter()
    for a in assets:
        m, cls = a.get("m") or "", a.get("cls") or ""
        if not m: continue
        canon = CANON.get(m.upper(), {}).get(cls) or ALIAS.get(m)
        for k in keys(canon or m): reg[k].add((cls, m))
        counts[(cls, m)] += 1

    PURPLE, TEAL = "FFCC99FF", "FF00B0B0"
    models, unmatched, matched = {}, [], []
    for r in range(3, wv.max_row + 1):
        label = wv.cell(r, 2).value
        if not label or str(label).strip() in ("Fleet", "TOTAL"): continue
        label = spell(re.sub(r"\s+", " ", str(label)).strip())

        got = []
        for cp in comps:
            single = cp.get("single")
            cap  = wv.cell(r, cp["col"]).value
            # A singleton has no interval and no OEM column beside it - reading
            # cp["col"]+1 would pick up whatever the next component's capacity
            # happens to be and call it a change interval.
            freq = None if single else wv.cell(r, cp["col"] + 1).value
            oem  = None if single else wv.cell(r, cp["col"] + 2).value
            if cap in (None, 0, "") and not oem: continue
            fill = wf.cell(r, cp["col"]).fill
            rgb  = fill.fgColor.rgb if (fill and fill.fgColor) else None
            rgb  = rgb if isinstance(rgb, str) else None
            cm   = wf.cell(r, cp["col"]).comment

            c = {"k": cp["k"], "en": cp["en"], "ru": cp["ru"]}
            if isinstance(cap, (int, float)) and cap: c["cap"] = round(float(cap), 2)
            if isinstance(freq, (int, float)) and freq: c["iv"] = int(freq)
            if oem: c["oem"] = re.sub(r"\s+", " ", str(oem)).strip()
            ty = cp.get("t") or LEGEND_TYPE.get(cp["k"])
            if ty: c["t"] = ty
            # The site's own flags, carried across as data instead of as a fill
            # colour nobody outside Excel can see.
            if rgb == PURPLE or (cm and "VERIFY" in str(cm.text).upper()):
                c["verify"] = 1
            if rgb == TEAL or ("cap" in c and "iv" not in c):
                c["noiv"] = 1
            if str(c.get("oem", "")).upper() == "VERIFY":
                c["verify"] = 1
            if cp["k"] in QUESTION_TYPES and c.get("oem"):
                c["ask"] = 1
            rx = CONTRADICTS.get(c.get("t"))
            if rx and c.get("oem") and rx.search(c["oem"]):
                c["ask"] = 1
            got.append(c)
        if not got: continue

        # No capacity and no interval, VERIFY set, so it lands on the work list
        # as figures to confirm rather than as a fact nobody supplied.
        if any(c["k"] in DRILL_CODES for c in got) and \
           not any(c["k"] == "7" for c in got):
            got.append(dict(ROTARY_HEAD))
            rotary.append(label)

        hits = set()
        for k in keys(label): hits |= reg.get(k, set())
        seen = sorted(hits)
        if seen: matched.append((label, sorted({r for _, r in seen})))
        if not seen:
            unmatched.append(label)
            models["?|" + label] = {"cls": "?", "m": label, "n": 0, "regs": [], "comps": got,
                                    "sourced": sum(1 for c in got if "cap" in c)}
            continue
        # ONE entry per class per machine, keyed by the MASTERLIST's name.
        # The register spells the same truck three ways — "KOMATSU", "KOMATSU
        # HM400", and nothing at all — and keeping one entry per spelling means
        # the canonical name resolves to two rivals and therefore to neither,
        # while the unit count is split across them. They are one machine, so
        # they are one row, and every spelling is an alias onto it.
        by_cls = collections.defaultdict(list)
        for (cls, regname) in seen: by_cls[cls].append(regname)
        for cls, regnames in by_cls.items():
            key = cls + "|" + label
            n = sum(counts[(cls, rn)] for rn in regnames)
            models[key] = {"cls": cls, "m": label, "regs": sorted(regnames),
                           "n": n, "comps": got,
                           "sourced": sum(1 for c in got if "cap" in c)}

    # ── what the import could not answer ─────────────────────────────────
    oems = collections.Counter()
    verify = noiv = ask = 0
    for M in models.values():
        for c in M["comps"]:
            if c.get("oem"): oems[c["oem"]] += 1
            verify += c.get("verify", 0); noiv += c.get("noiv", 0); ask += c.get("ask", 0)

    used_codes = {c["k"] for m in models.values() for c in m["comps"]}
    unused_codes = sorted(k for k in legend if k not in used_codes)

    with open(REPORT, "w", encoding="utf-8") as f:
        f.write("LUBRICATION MASTERLIST — IMPORT REPORT\n")
        f.write("=" * 62 + "\n\n")
        f.write("Generated by docs/build-lube-data.py. Everything here is a question\n")
        f.write("for the reliability engineer, not a fault in the import.\n\n")
        f.write("models imported            %5d\n" % len(models))
        f.write("compartment entries        %5d\n" % sum(len(m["comps"]) for m in models.values()))
        f.write("distinct OEM spec strings  %5d   <- the standardisation problem\n" % len(oems))
        f.write("flagged VERIFY             %5d\n" % verify)
        f.write("missing change interval    %5d\n" % noiv)
        f.write("OEM contradicts the type  %5d\n\n" % ask)

        f.write("COMPONENT NAMES REPAIRED FROM THE LEGEND (%d)\n" % len(renamed)
                + "-" * 62 + "\n")
        f.write("The header row of the frequency sheet had Russian in the English\n"
                "field for these, one truncated word, and one with a space broken\n"
                "into the middle of a word. The Legend tab has had a proper English\n"
                "name for every one of them all along and nobody was reading it.\n\n")
        for k, was, now in renamed:
            f.write("  %-5s %-42s -> %s\n" % (k, was[:42], now))

        f.write("\nNAMES THE SHEET AND THE LEGEND DISAGREE ON (%d)\n" % len(differ)
                + "-" * 62 + "\n")
        f.write("Kept as the sheet has them. Neither is broken, so neither is\n"
                "rewritten - but two names for one component is two names on two\n"
                "pieces of paper, and somebody should pick one.\n\n")
        for k, sheet_name, leg in differ:
            f.write("  %-5s sheet: %-36s legend: %s\n" % (k, sheet_name[:36], leg))

        f.write("\nROTARY HEAD ADDED, FIGURES STILL MISSING (%d)\n" % len(rotary)
                + "-" * 62 + "\n")
        f.write("Code 7 is in the Legend and the frequency sheet has no column\n"
                "for it. These are drill rigs, so the compartment is real: it is\n"
                "carried with no capacity and no interval, flagged to confirm.\n"
                "Derived from the drill-rig codes the masterlist already gives a\n"
                "model (%s), so it grows with the sheet -\n"
                "a drill whose row carries none of those gets no rotary head and\n"
                "is worth a look.\n\n" % ", ".join(sorted(DRILL_CODES)))
        for m in rotary:
            f.write("  %s\n" % m)

        f.write("\nCODES THE LEGEND DEFINES THAT NO MACHINE USES (%d)\n"
                % len(unused_codes) + "-" * 62 + "\n")
        f.write("Either the fleet genuinely has none, or a column is missing from\n"
                "the frequency sheet. Code 10 (fuel) is deliberate - the Legend\n"
                "itself says fuel is not a lubricant.\n\n")
        for k in unused_codes:
            f.write("  %-5s %s\n" % (k, legend.get(k, "")))
        f.write("\n")

        f.write("MODELS IN THE MASTERLIST WITH NO MACHINE ON THE REGISTER (%d)\n" % len(unmatched))
        f.write("They are imported and carry no unit count, so they cost nothing\n")
        f.write("and appear the moment a matching machine is registered.\n")
        for u in sorted(unmatched): f.write("   " + u + "\n")
        f.write("\nEXPLICIT ALIASES — every one a judgement, check them (%d)\n" % len(ALIAS))
        for k, v in sorted(ALIAS.items()):
            f.write("   register %-28s = masterlist %s\n" % (k, v))
        f.write("\nNAMES MATCHED ACROSS THE TWO DOCUMENTS (%d)\n" % len(matched))
        f.write("Every one of these is a judgement the importer made. Check them.\n")
        for lab, regs in sorted(matched):
            if [lab] != regs:
                f.write("   %-42s = %s\n" % (lab, ", ".join(regs)))
        f.write("\nOEM SPEC STRINGS, MOST USED FIRST\n")
        for k, v in oems.most_common(): f.write("  %4d  %s\n" % (v, k))

    HEAD = '''/* Lubrication reference — GENERATED from the site's own masterlist.

   Source: docs/source/Lube_Matrix_Oil_Analysis_Sampling.xlsm, sheet
   "Lube Frequency BMSK". Rebuild with docs/build-lube-data.py.
   Import questions: docs/lube-import-report.txt

   THE CODES AND THE FIGURES ARE THE CLIENT'S. The component codes (1, 2, 3,
   3A, 4AL … 16) are on the printed sampling forms and in the fitters' heads;
   the capacities and intervals came out of OEM manuals. Nothing here is
   invented, and where the masterlist contradicts itself the contradiction is
   carried through and flagged rather than quietly resolved.

   FLAGS, which were cell colours in Excel and are data here:
     verify  the figure is a placeholder to confirm against the manual (purple)
     noiv    a capacity with no change interval, so it totals nothing (teal)
     ask     the Legend files this under gear oil but the OEM code reads TO-4;
             wet-clutch friction chemistry is an engineer's decision

   THE FIELD NEVER SEES `oem`. It is 91 different strings for eight products —
   Japanese full-width, Russian, brand names, multi-line — and asking a fitter
   in gloves to read it is how the wrong oil goes in. It is kept because it is
   what the standard is DECIDED against, on the dashboard, once. */
(function (G) {
'''

    TAIL = '''
  /* Colour is by LUBRICANT TYPE, which is the site's own convention: blue
     engine, red hydraulic, yellow compressor, green rock-drill. Eight colours
     to learn, and they survive a change of supplier — colouring by product
     would have to be relearned the day a drum changes. */
  function typeOf(k){ return (COMP_TYPE[k] || null); }
  function productFor(k){
    var t = typeOf(k); if(!t) return null;
    return CATALOG.filter(function(p){ return p.t === t; })[0] || null;
  }
  function norm(s){ return String(s||"").toLowerCase().replace(/[^a-z0-9]+/g," ").trim(); }
  var REG = null;                      /* register(), built once — see below */
  function resolve(model, cls){
    if(!model) return null;
    if(cls && MODELS[cls + "|" + model]) return cls + "|" + model;
    var n = norm(model), hits = [];
    Object.keys(MODELS).forEach(function(k){
      var r = MODELS[k];
      /* The canonical name, or any spelling the register uses for it. */
      var names = [r.m].concat(r.regs || []);
      if(!names.some(function(x){ return norm(x) === n; })) return;
      if(cls && r.cls !== cls) return;
      hits.push(k);
    });
    return hits.length === 1 ? hits[0] : null;
  }

  G.LUBE = {
    models:   Object.keys(MODELS),
    of:       function(model, cls){ var k = resolve(model, cls); return k ? MODELS[k] : null; },
    ambiguous:function(model){
                var n = norm(model), c = 0;
                Object.keys(MODELS).forEach(function(k){
                  var r = MODELS[k];
                  if([r.m].concat(r.regs || []).some(function(x){ return norm(x) === n; })) c++;
                });
                return c > 1;
              },
    comps:    function(model, cls){ var r = this.of(model, cls); return r ? r.comps : []; },
    comp:     function(model, k, cls){
                return this.comps(model, cls).filter(function(c){ return c.k === k; })[0] || null; },
    label:    function(model, k, lang, cls){
                var c = this.comp(model, k, cls);
                return c ? (lang === "ru" ? c.ru : c.en) : k; },
    /* Sourced means a capacity that is not a placeholder. A purple cell in the
       masterlist is a typical figure somebody filled in to make the totals
       work, and counting it as known is how a guess becomes a fact. */
    sourced:  function(model, k, cls){
                var c = this.comp(model, k, cls);
                return !!(c && c.cap != null && !c.verify); },
    catalog:  CATALOG,
    product:  function(name){
                return CATALOG.filter(function(p){ return norm(p.p) === norm(name); })[0] || null; },
    /* What belongs in this compartment: one product, by type. The whole point
       of the exercise — the fitter is never offered a choice of eight. */
    forComp:  function(model, k, cls){
                var c = this.comp(model, k, cls);
                return c ? productFor(c.k) : null; },
    typeOf:   typeOf,
    types:    TYPES,
    hue:      function(t){ return (TYPES[t] && TYPES[t].hue) || "#8b969c"; },
    brands:   BRANDS,
    brandHue: function(b){ return (BRANDS[b] && BRANDS[b].hue) || null; },
    /* Brand is shown only where it still MATTERS — that is, where more than one
       brand serves the same job. The point of standardising is to end with one
       brand per lubricant type, so a permanent brand mark teaches the mess. Let
       it fade as the site converges and the edge becomes the gap, visible. */
    brandsFor: function(type){
      var out = {};
      CATALOG.forEach(function(p){ if(p.t === type && p.brand) out[p.brand] = 1; });
      return Object.keys(out);
    },
    brandMatters: function(type){ return this.brandsFor(type).length > 1; },
    site:     SITE,
    /* Everything the masterlist could not answer, as a work list rather than a
       cell colour: figures to confirm, intervals that are missing, and the
       transmission-oil question. */
    gaps:     function(){
                var out = { verify:[], noiv:[], ask:[] };
                Object.keys(MODELS).forEach(function(k){
                  MODELS[k].comps.forEach(function(c){
                    ["verify","noiv","ask"].forEach(function(f){
                      if(c[f]) out[f].push({ key:k, m:MODELS[k].m, k:c.k,
                                             en:c.en, cap:c.cap, oem:c.oem });
                    });
                  });
                });
                return out;
              },
    EVID: [
      { k:"label", rank:2, en:"Label photographed", ru:"Фото этикетки" },
      { k:"batch", rank:2, en:"Tank / pump batch",  ru:"Партия из ёмкости" },
      { k:"told",  rank:1, en:"Reported, not shown",ru:"Со слов" }
    ],
    evidRank: function(k){
      var e = this.EVID.filter(function(x){ return x.k === k; })[0];
      return e ? e.rank : 0;
    },

    /* ---- what is on the property, and whether it belongs where it is -------

       Both lists as one, deduplicated by name: the EXSOIL products in the
       machines today and the Lemarc/Teboil ones bought for 2027. `src` says
       which list a product came from, "both" when it is on each of them.

       Built once. It is asked per compartment, and a lubrication round on a
       haul truck asks eighteen times. */
    register: function(){
      if (REG) return REG;
      var out = [], at = {};
      var add = function(p, t, src){
        if (!p) return;
        var k = norm(p);
        if (at[k]) { if (at[k].src !== src) at[k].src = "both"; return; }
        at[k] = { p: p, t: t || "", src: src };
        out.push(at[k]);
      };
      CATALOG.forEach(function(x){ add(x.p, x.t, "field"); });
      ((G.LUBE2027 && G.LUBE2027.shelf) || []).forEach(function(x){ add(x.p, x.t, "2027"); });
      return (REG = out);
    },
    registered: function(name){
      var n = norm(name);
      return this.register().filter(function(x){ return norm(x.p) === n; })[0] || null;
    },

    /* ONE judgement of one compartment, for every screen that asks.

       It used to be three: the capture screen judged by lubricant TYPE, the
       dashboard's position card compared product names as strings, and the
       report had a fourth rule of its own. The site is changing supplier for
       2027, so name equality calls a correctly-filled compartment wrong the
       day the new drums land - which is what the dashboard did while the phone
       standing at the same machine said it conformed.

       What destroys a machine is the wrong KIND of oil: engine oil in a final
       drive, GL-5 in a wet brake. A product that serves this compartment's job
       is right whichever supplier it came from; one that serves a different job
       is the finding; one on neither list is unjudged and says so. */
    verdict: function(model, cls, key, name){
      var nm = String(name == null ? "" : name).trim();
      if (!nm) return { b:"none", k:"lube_v_none", want:"", product:"" };
      var want = this.forComp(model, key, cls);
      /* Wire rope and open gear are both still "(verify product)" on the
         site's own legend. Recording what is in there is worth doing;
         pretending it can be judged is not. */
      if (!want) return { b:"none", k:"lube_v_nostd", want:"", product:nm };
      var c = this.comp(model, key, cls), ty = (c && c.t) || want.t || "";
      var prod = this.registered(nm);
      var r = { want: want.p || "", product: nm };
      /* Typed in by hand: a real finding - somebody put something in that is on
         neither list - and it must not be dressed up as a pass or a failure. */
      if (!prod)                     { r.b = "watch"; r.k = "lube_v_unknown"; }
      else if (ty && prod.t && prod.t !== ty) { r.b = "act"; r.k = "lube_v_wrong"; }
      else if (!ty || !prod.t)       { r.b = "watch"; r.k = "lube_v_unknown"; }
      else                           { r.b = "ok";    r.k = "lube_v_ok"; }
      /* "Holds something other than the site standard" - the count the report
         puts above the table. Anything not judged conforming belongs in it;
         an unlisted product is exactly the case somebody has to go and look at. */
      r.off = r.b !== "ok";
      return r;
    }
  };
})(window);
'''

    comp_type = {k: v for k, v in LEGEND_TYPE.items() if v}
    types = {}
    for p in products:
        types[p["t"]] = {"en": p["en"], "ru": p["ru"], "hue": p["hue"]}
    types.setdefault("wirerope", {"en":"Wire rope lube","ru":"Смазка канатов","hue":"#5b686f"})
    types.setdefault("opengear", {"en":"Open gear grease","ru":"Смазка открытых передач","hue":"#8c5a2b"})
    brands = {b: {"hue": h} for b, h in BRAND_HUE.items()}

    with open(OUT, "w", encoding="utf-8") as f:
        f.write(HEAD)
        f.write("  var SITE = " + json.dumps(
            {"design": -40, "winter": -45, "summer": 28, "hoursPerYear": 5000,
             "units": len(assets)}) + ";\n\n")
        f.write("  /* The eight products actually on site, from row 1 of the masterlist. */\n")
        f.write("  var CATALOG = " + json.dumps(products, ensure_ascii=False) + ";\n\n")
        f.write("  var TYPES = " + json.dumps(types, ensure_ascii=False) + ";\n\n")
        f.write("  /* Brand is the EDGE, not the fill — see build-lube-data.py. */\n")
        f.write("  var BRANDS = " + json.dumps(brands, ensure_ascii=False) + ";\n\n")
        f.write("  var COMP_TYPE = " + json.dumps(comp_type, ensure_ascii=False) + ";\n\n")
        f.write("  var MODELS = " + json.dumps(models, ensure_ascii=False,
                                               separators=(",", ":")) + ";\n")
        f.write(TAIL)

    print("models          %5d" % len(models))
    print("entries         %5d" % sum(len(m["comps"]) for m in models.values()))
    print("products        %5d" % len(products))
    print("OEM strings     %5d" % len(oems))
    print("VERIFY flags    %5d" % verify)
    print("missing interval%5d" % noiv)
    print("unmatched models%5d  (see docs/lube-import-report.txt)" % len(unmatched))
    print("bytes           %5d" % len(open(OUT, encoding="utf-8").read()))

main()
